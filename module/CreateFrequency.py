from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font


class CreaFrequency:
    def __init__(self, date):
        self._date = date
        self._excel = Workbook()

    def header(self, sheet, title='AGENTE DE FISCALIZAÇÃO'):
        area = sheet
        if sheet in ['COORDENADOR GERAL', 'COORDENADOR', 'SUPERVISOR', 'INTERNO', 'AGENTE DE FISCALIZAÇÃO']:
            area = 'INTERNOS'
            title = sheet

        frequency = self._excel[sheet]

        font14 = Font(b=True, size=14)
        font11 = Font(b=True, size=11)
        alignment = Alignment(horizontal='center',
                              vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        logo_salvador = Image('./assets/logo_salvador.png')
        logo_transalvador = Image('./assets/logo_transalvador.png')
        frequency.add_image(logo_salvador, 'A1')
        frequency.add_image(logo_transalvador, 'G1')

        frequency.column_dimensions['A'].width = 3
        frequency.column_dimensions['B'].width = 30
        frequency.column_dimensions['C'].width = 10
        frequency.column_dimensions['D'].width = 30
        frequency.column_dimensions['E'].width = 10
        frequency.column_dimensions['F'].width = 10
        frequency.column_dimensions['G'].width = 10
        frequency.column_dimensions['H'].width = 10
        frequency.column_dimensions['I'].width = 10

        frequency.row_dimensions[1].height = 55
        frequency.row_dimensions[2].height = 35
        frequency.row_dimensions[4].height = 35

        frequency.merge_cells('A1:I1')  # 'images'
        frequency.merge_cells('A2:I2')  # 'LISTA DE FREQUENCIA'
        frequency.merge_cells('A3:C3')  # space area
        frequency.merge_cells('E3:I3')  # space date
        frequency.merge_cells('A4:I4')  # space title

        #[nome, matricula, assinatura]
        frequency.merge_cells('A5:B6')
        frequency.merge_cells('C5:C6')
        frequency.merge_cells('D5:D6')
        frequency.merge_cells('E5:F5')
        frequency.merge_cells('G5:H5')
        frequency.merge_cells('I5:I6')

        for cell in ['E3', 'A3','A4', 'A2', 'A5', 'C5', 'D5', 'E5', 'G5', 'I5', 'E6', 'F6', 'G6', 'H6']:
            frequency[cell].font = font11
            frequency[cell].alignment = alignment
            if int(cell[1]) > 4:
                frequency[cell].border = border

        frequency['E3'] = self._date
        frequency['A3'] = area.upper()
        frequency['A4'] = title
        frequency['A2'] = 'LISTA DE FREQUÊNCIA'
        frequency['A5'] = 'Nome'
        frequency['C5'] = 'Matrícula'
        frequency['D5'] = 'ASSINATURA'
        frequency['E5'] = '1º TURNO'
        frequency['G5'] = '2º TURNO'
        frequency['I5'] = 'QTD HORAS'
        frequency['E6'] = 'Entrada'
        frequency['F6'] = 'Saída'
        frequency['G6'] = 'Entrada'
        frequency['H6'] = 'Saída'

        frequency['A2'].font = font14
        frequency['A4'].font = font14

    def create_frequency(self, sheet, workers, position=7):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        frequency = self._excel.create_sheet(sheet)
        self.header(sheet)
        line_begin = position
        for i, worker in enumerate(workers, 1):
            for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                frequency[f'{c}{line_begin}'].border = border

            frequency[f'A{line_begin}'] = i
            frequency[f'B{line_begin}'] = worker[0]
            frequency[f'C{line_begin}'] = worker[1]
            line_begin += 1
        return line_begin

    def create_changes(self, sheet, changes, position):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        frequency = self._excel[sheet]

        line_begin = position
        for i, change in enumerate(changes, 1):
            for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                frequency[f'{c}{line_begin}'].border = border

            frequency[f'A{line_begin}'] = i
            frequency[f'B{line_begin}'] = change[0]
            frequency[f'C{line_begin}'] = change[1]
            #frequency[f'D{line_begin}'] = change[2]
            line_begin += 1

    def formatCellChange(self, sheet, last_position):
        font11 = Font(b=True, size=11)
        alignment = Alignment(horizontal='center',
                              vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        frequency = self._excel[sheet]
        frequency.merge_cells(f'A{last_position}:I{last_position}')
        frequency[f'A{last_position}'] = 'PERMUTAS  - SUBSTITUIÇÕES - EMERGENCIAIS'
        frequency[f'A{last_position}'].border = border
        frequency[f'A{last_position}'].font = font11
        frequency[f'A{last_position}'].alignment = alignment



    def save(self, path='./'):
        self._excel.save(f'{path}frequencia_{self._date}.xlsx')

    def run(self, frequencys, path='./'):
        for area in frequencys.keys():
            last_position = self.create_frequency(area, frequencys[area]['workers'])
            self.formatCellChange(area, last_position)
            self.create_changes(area, frequencys[area]['changes'], last_position+1)

        self.save(path)


if __name__ == '__main__':

    frequencys = {
        'area1': 
            {
                'workers': [
                    ('Friedmann Keller Pereira da Silva', '3067479'),
                    ('Antonio Marco Teixeira Souza', '3067340'),
                    ('Gilson da Conceição Machado', '3097959'),
                    ('Isabel Valente Tinel', '0000000')
                ],
                'changes':[
                    ('Isabel Valente Tinel', '3165395','Diego Galvão Carvalho')
                ]
            },
    }

    CreaFrequency('2022-11-11').run(frequencys, './assets/')
